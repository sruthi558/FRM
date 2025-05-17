from flask import Flask, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_jwt_extended import (JWTManager, create_access_token, decode_token,jwt_required, get_jwt_identity)
from flask_cors import CORS
from flask_mail import Mail, Message
import email
import imaplib
from email.header import decode_header
from openpyxl import load_workbook
from openpyxl import workbook

from threading import Thread
import psycopg2
from psycopg2.extras import DictCursor
import pandas as pd
import os
import schedule
import time
import threading
from apscheduler.schedulers.background import BackgroundScheduler
import base64
import atexit
from datetime import date,datetime,timedelta
from flask import current_app as app
from apscheduler.events import EVENT_JOB_EXECUTED, EVENT_JOB_ERROR
import re
import mail
from io import BytesIO  
from openpyxl import load_workbook,Workbook

from flask import Flask, send_file
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.styles.colors import COLOR_INDEX, Color
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.utils import range_boundaries
import traceback


import os
from dotenv import load_dotenv

load_dotenv()



app = Flask(__name__)
CORS(app)

app.config['SQLALCHEMY_DATABASE_URI'] = f"postgresql://{os.getenv('DB_USER')}:{os.getenv('DB_PASSWORD')}@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = 'your_jwt_secret_key'
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'durga@corundumtechnologies.com'
app.config['MAIL_PASSWORD'] = 'yaviezhynflcygvs'
app.config['MAIL_DEFAULT_SENDER'] = 'your-email@gmail.com'
app.config['IMAP_SERVER'] = 'imap.gmail.com'
app.config['IMAP_PORT'] = 993
app.config['EMAIL_ACCOUNT'] = 'durga@corundumtechnologies.com'
app.config['EMAIL_PASSWORD'] = 'yaviezhynflcygvs'
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

db = SQLAlchemy(app)
jwt = JWTManager(app)
mail = Mail(app)
dbname = os.getenv('DB_NAME')
user = os.getenv('DB_USER')
password = os.getenv('DB_PASSWORD')
host = os.getenv('DB_HOST')
port = os.getenv('DB_PORT')
app.config['UPLOAD_FOLDER'] = 'uploads/'
class UserDetails(db.Model):
    __tablename__ = 'user_details'
    id = db.Column(db.Integer, primary_key=True)
    emp_id = db.Column(db.String(255), nullable=False)
    username = db.Column(db.String(255), unique=True, nullable=False)
    email_id = db.Column(db.String(255), unique=True, nullable=False)
    mobile_no = db.Column(db.String(15), nullable=False)
    address = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(25), nullable=False)   
    telephone_no = db.Column(db.String(20), nullable=False)
    branch = db.Column(db.String(100), nullable=False)
    password = db.Column(db.String(255), nullable=True)     
    is_admin = db.Column(db.Boolean, default=False)
    branch_region = db.Column(db.String(100), nullable=False)
    def __repr__(self):
        return f'<UserDetails {self.username}>'
class OutputTable(db.Model):
    __tablename__ = 'output_table'
    unique_id = db.Column(db.Integer, primary_key=True)
    comments_by_ro_maker = db.Column(db.String(500))
    date_updated_by_ro_maker = db.Column(db.DateTime)
    comments_by_ro_checker = db.Column(db.String(500))
    date_updated_by_ro_checker = db.Column(db.DateTime)
    comments_by_co_maker=db.Column(db.String(500))
    date_updated_by_co_maker = db.Column(db.DateTime)
    co_maker_status = db.Column(db.String(1000))
    co_maker_action= db.Column(db.String(1000))
    co_maker_scn_status = db.Column(db.String(500))
    co_maker_scn_comments = db.Column(db.String(1000))
    co_maker_scn_date_updated = db.Column(db.DateTime)
    comments_by_co_checker=db.Column(db.String(500))
    date_updated_by_co_checker = db.Column(db.DateTime)
    co_checker_status=db.Column(db.String(1000))
    co_checker_action=db.Column(db.String(1000))
    co_checker_scn_action = db.Column(db.String(500))
    co_checker_scn_comments = db.Column(db.String(1000))
    co_checker_scn_date_updated = db.Column(db.DateTime)
    reply_body = db.Column(db.Text)
    ro_maker_files=db.Column(db.JSON)
    ro_checker_files=db.Column(db.JSON)
    co_maker_files = db.Column(db.JSON)
    co_checker_files=db.Column(db.JSON)
    attachments = db.Column(db.JSON)
    distribute = db.Column(db.Boolean, default=False)
    frm_cell_head=db.Column(db.String(500))
    frm_cell_head_files=db.Column(db.JSON)
    date_updated_by_frm_cell_head= db.Column(db.DateTime)
    audit_comment =db.Column(db.String(500))
    date_updated_by_audit = db.Column(db.DateTime)
    audit_files=db.Column(db.JSON)
    frm_cell_head_scn_comments=db.Column(db.String(500))
    frm_cell_head_scn_date_updated=db.Column(db.DateTime)
    fraud_cat_report_rbi= db.Column(db.String(500))
    frm_after_audit=db.Column(db.String(500))
    frm_after_audit_files=db.Column(db.JSON)
    frm_after_audit_date=db.Column(db.DateTime)
    frm_cell_head_sendback_reason= db.Column(db.String(500))
    co_checker_sendback_reason= db.Column(db.String(500))
    co_maker_sendback_reason= db.Column(db.String(500))
    ro_checker_sendback_reason=db.Column(db.String(500))
    sendback_to=db.Column(db.String(500))
    submited_to= db.Column(db.String(500))
    def __repr__(self):
        return f"<OutputTable {self.unique_id}>"


@app.route('/fetch_matched_data_from_output_table')
@jwt_required()
def fetch_matched_data_from_output_table():
    cursor = None
    connection = None           
    print('1')
    try:
        user_id = get_jwt_identity()
        print("22")
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        user_role = user.role
        print(user_role,"user")
        print('2')

        user_branch_region=user.branch_region
        connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),  
            password=os.getenv('DB_PASSWORD'),
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        
        cursor = connection.cursor()
        print("Connected to the database.")
        cursor.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'output_table' AND column_name = 'distribute';   
        """)                                                                                                           
        column_check = cursor.fetchone()                                                           

        if not column_check:
            cursor.execute("""
                ALTER TABLE output_table 
                ADD COLUMN distribute BOOLEAN DEFAULT FALSE;
            """)
            connection.commit()
            print("Column 'distribute' added to 'output_table' with default value False.")

        distribute_condition = "o.distribute IS TRUE"
        distribute_condition1 = "o.distribute IS  False"

        if user_role == 'Officer1':
             query = f"""
    SELECT DISTINCT ON(o.unique_id)
        o.unique_id,  -- Assuming this is the unique identifier for the alert
        o.account_number AS acctno,     -- Account number from output_table
        o.customer_id AS custcd,        -- Customer ID from output_table
        o.query_number AS rule_number,  -- Rule number (query_no from output table)
        o.description,                  -- Description from output table
        o.date AS raised_date,          -- Date (Alert raised date) from output table
        o.branch_region                 -- Branch region from output table
    FROM output_table o
    WHERE {distribute_condition1} ;
    """
        elif user_role == 'RO Maker' and user_branch_region:
            query = f"""
            SELECT DISTINCT ON(o.unique_id)
                o.unique_id,  -- Assuming this is the unique identifier for the alert
                c.acctno,     -- Account number from customers table
                c.custcd,     -- Customer ID from customers table
                c.customername,  -- Customer name from customers table
                o.branch,
                
                o.query_number AS rule_number,  -- Rule number (query_no from output table)
                o.description,  -- Description from output table
                o.date AS raised_date,  -- Date (Alert raised date) from output table
                o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd  
            WHERE  o.branch_region = '{user_branch_region}' AND 
            {distribute_condition}
             AND
            o.comments_by_ro_maker IS NULL ;
            """
        elif user_role == 'RO Checker' and user_branch_region:
            query = f"""
            SELECT DISTINCT ON(o.unique_id)
                o.unique_id,  -- Assuming this is the unique identifier for the alert
                c.acctno,     -- Account number from customers table
                c.custcd,     -- Customer ID from customers table
                c.customername,  -- Customer name from customers table
                o.branch,
                o.query_number AS rule_number,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,
                o.ro_maker_files,
                o.description,  -- Description from output table
                o.date AS raised_date,  -- Date (Alert raised date) from output table
                o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.comments_by_ro_maker IS NOT NULL  AND (o.comments_by_ro_checker IS NULL OR o.comments_by_ro_checker IS NOT NULL) AND
            o.branch_region = '{user_branch_region}' AND 
            {distribute_condition} AND o.submited_to = 'RO Checker' AND (o.sendback_to='RO Checker' OR o.sendback_to IS Null)
            ;
            """
        elif user_role == 'CO Maker':
            query = """
            SELECT DISTINCT ON(o.unique_id)
        o.unique_id,  -- Assuming this is the unique identifier for the alert
        c.acctno,     -- Account number from customers table
        c.custcd,     -- Customer ID from customers table
        c.customername,  -- Customer name from customers table
        o.branch,
        o.query_number AS rule_number,
        o.comments_by_ro_checker,
        o.ro_maker_files,
        o.ro_checker_files,
        o.date_updated_by_ro_checker,
        o.comments_by_ro_maker,
        o.date_updated_by_ro_maker,
        
        o.description,  -- Description from output table
        o.date AS raised_date,  -- Date (Alert raised date) from output table
        o.branch_region  -- Branch region from output table
    FROM output_table o
    JOIN user_details u ON o.branch_region = u.branch_region
    LEFT JOIN customers c ON o.customer_id = c.custcd
    WHERE o.comments_by_ro_checker IS NOT NULL  AND (o.comments_by_co_maker IS NULL OR o.comments_by_co_maker IS NOT NULL) AND o.submited_to = 'CO Maker' AND (o.sendback_to='CO Maker' OR o.sendback_to IS Null);
    """
            
        elif user_role == 'CO Checker':
          print('fetch co checker')
          query = """
           SELECT DISTINCT ON (o.unique_id)
                o.unique_id,  -- Assuming you want distinct results based on the unique_id
                c.acctno,     
                c.custcd, 
                c.customername, 
                o.branch,
                o.comments_by_ro_checker,
                o.ro_maker_files,
                o.ro_checker_files,
                o.date_updated_by_ro_checker,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,
                o.query_number AS rule_number,  
                o.description,    
                o.*,              
                o.date AS raised_date,          
                o.branch_region,
                u.*, c.*  -- or you can select specific columns from user_details and customers if needed
                FROM output_table o
                JOIN user_details u ON o.branch_region = u.branch_region
                LEFT JOIN customers c ON o.customer_id = c.custcd
                WHERE o.comments_by_co_maker IS NOT NULL AND (o.comments_by_co_checker IS NULL OR o.comments_by_co_checker IS NOT NULL) AND o.submited_to = 'CO Checker'
                AND (o.sendback_to='CO Checker' OR o.sendback_to IS Null); 
            """  
        elif user_role == 'FRM Cell Head':
            query ="""
       SELECT DISTINCT ON (o.unique_id) 
            o.unique_id,
            o.*, 
            u.*, 
            c.*,
            o.email_id, 
            o.query_number AS rule_number,  
            o.date AS raised_date,
            o.customer_id AS custcd
        FROM output_table o
        JOIN user_details u ON o.branch_region = u.branch_region
        LEFT JOIN customers c ON o.customer_id = c.custcd
        WHERE 
         o.comments_by_co_checker IS NOT NULL AND (o.frm_cell_head IS NULL OR o.frm_cell_head IS NOT NULL)
        AND o.co_checker_status IN ('Recommended For making as RFA', 'Under Investigation')
        AND o.submited_to = 'FRM Cell Head' ; 
        """

        elif user_role == 'Audit':
            query = """
            SELECT DISTINCT ON (o.unique_id) 
                o.*, 
                u.*, 
                c.*,
                o.email_id, 
                o.date AS raised_date,
                o.customer_id AS custcd,
                o.query_number AS rule_number,
                o.ro_maker_files,
                o.ro_checker_files,
                o.co_maker_files,
                o.co_checker_files,
                o.frm_cell_head_files
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.frm_cell_head IS NOT NULL AND o.audit_comment IS NULL;
            """
        else:
            query = """
            SELECT DISTINCT o.*
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region;
            """
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]
        for row in matched_data:
            if 'raised_date' in row and row['raised_date']:
                if isinstance(row['raised_date'], datetime):
                    row['raised_date'] = row['raised_date'].strftime('%Y-%m-%d')
                else:
                    try:
                        row['raised_date'] = datetime.strptime(row['raised_date'], '%Y-%m-%d').strftime('%Y-%m-%d')
                    except Exception as e:
                        print(f"Error parsing date: {e}")
        if 'date_updated_by_frm_cell_head' in columns:
            updated_dates = [row.get('date_updated_by_frm_cell_head') for row in matched_data]
            print("Date Updated by FRM Cell Head:", updated_dates)
        else:
            print("The column 'date_updated_by_frm_cell_head' was not found in the query results.")


        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")


@app.route('/update_comment/<int:unique_id>', methods=['POST'])
@jwt_required()
def update_comment(unique_id):
    token = request.headers.get('Authorization').split()[1]
    decoded_token = decode_token(token)
    data = request.get_json()
    comment = data.get('comment')
    print(comment,"=================")
    user_id = get_jwt_identity()
    user = UserDetails.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    user_role = user.role
    record = OutputTable.query.filter_by(unique_id=unique_id).first()
    if not record:
        return jsonify({"message": "Record not found"}), 404
    if user_role == 'RO Checker':
        record.comments_by_ro_checker = comment
        print("comments updated")
    elif user_role == 'RO Maker':
        record.comments_by_ro_maker = comment
        print("comments updated")
    elif user_role == 'CO Maker':
        record.comments_by_co_maker = comment
        print("comments updated")
    elif user_role == 'CO Checker':
        record.comments_by_co_checker = comment
        print("comments updated")
    elif user_role == 'FRM Cell Head':
        record.frm_cell_head = comment
        record.frm_after_audit = comment
    elif user_role == 'Audit':
        record.audit_comment = comment
    else:
        return jsonify({"message": "Unauthorized role to update comments"}), 403
    db.session.commit()
    return jsonify({"message": "Comment updated successfully"}), 200

@app.route('/send_email', methods=['POST'])
def send_email():
    data = request.get_json()
    print(data,"data")
    recipient_email = data.get('email')      # Get the email address
    unique_id = data.get('alertno')             # Get the unique ID (loan account number)
    comment = data.get('frm_cell_head')           # Get the frmcellhead (comment)
    accno = data.get('account_number') # Get the loan account number from request
    custname=data.get('customername')   
    date=data.get('cocheckersubmiteddate')    
    print('customer name:',custname)    
    print('date:',date) 
    print('uni id:', unique_id)
    print('emailll:', recipient_email)
    print('commenting:', comment)
    print('loan account:', accno)


    if not recipient_email or recipient_email.lower() == 'nan' or recipient_email == "":
        if unique_id:
            # Update the database if no email is provided
            customer_data = OutputTable.query.filter_by(unique_id=unique_id).first()
            if customer_data:
                customer_data.frm_cell_head = comment  # Save the comment to the database
                db.session.commit()
                print(f"Updated frm_cell_head for unique_id {unique_id}")
                return jsonify({'success': True, 'message': 'Comment updated but no valid email found (email is empty or NaN).'}), 200
            else:
                return jsonify({'success': False, 'message': 'Customer not found in database'}), 404
        else:
            return jsonify({'success': False, 'message': 'No unique_id or email provided'}), 400

    try:
        # Prepare subject and body with the loan account number and the unique ID
        subject = f"Examination of irregularities observed in the Loan A/c {accno} from the Fraud angle, as per the guidelines of Reserve Bank of India - ID: {unique_id}"
        body = f"""
        Dear Sir/Madam,

        At your request, our Bank has sanctioned/ granted from time to time various credit facilities aggregating to Rs ____ in the above loan account. It is observed that you have failed to maintain the loan account/s as per the sanctioned terms and consequent to the default committed by you, the loan account became irregular/ NPA. As certain irregularities were observed in your loan account/s, it has been decided to appoint Forensic Auditor / Internal Auditor for further investigation.

        M/s {custname}, Forensic Auditor/ Internal Auditor appointed in the account, has submitted their report dated {date} (copy enclosed) wherein certain serious irregularities/ anomalies/ allegations/ commissions / omissions were pointed out by the Auditors, needs to be examined by the Competent Authority from the angle of fraud based on the findings of the Forensic Auditors/ Internal Auditor.

        Therefore we call upon you to submit your response / reply, if any on the said irregularities / anomalies / allegations / commissions / omissions as pointed out the Auditors in the enclosed report within Twenty One (21) days from the receipt of this letter to enable us to place your response before the appropriate committee for a suitable decision. In case Bank do not receive any reply from your end within the stipulated time period, it shall be construed that you do not have anything to say in this regard and accordingly the same will be placed before the appropriate committee for their decision on further action including identification of the loan account of M/s {custname} under fraud or otherwise, as per the extant guidelines as contained in the Master Directions of RBI on Frauds.

        Yours faithfully
        Bank Name

        Encl: - Copy of the Report of Forensic Auditor / Internal Auditor

        """

        msg = Message(
            subject=subject,
            recipients=[recipient_email],
            body=body
        )
        msg.sender = app.config['MAIL_DEFAULT_SENDER']
        mail.send(msg)
        print("Email sent successfully!")

        # If email is sent, update frm_cell_head status in the database to 'completed'
        if unique_id:
            customer_data = OutputTable.query.filter_by(unique_id=unique_id).first()
            if customer_data:
                customer_data.frm_cell_head = comment  # Update the status in the database
                db.session.commit()
                print(f"Updated frm_cell_head for unique_id {unique_id}")
            else:
                return jsonify({'success': False, 'message': 'Customer not found in database'}), 404
        else:
            return jsonify({'success': False, 'message': 'No unique_id provided'}), 400

        return jsonify({'success': True, 'message': 'Email sent successfully',"redirect": "/dashboard"}), 200
    except Exception as e:
        print(f"Error sending email: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500



@app.route('/submit_form/<int:unique_id>', methods=['POST']) 
@jwt_required()
def submit_form(unique_id):
    print('Received unique_id:', unique_id)
    token = request.headers.get('Authorization').split()[1]
    decoded_token = decode_token(token)
    user_id = get_jwt_identity()
    connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),  
            password=os.getenv('DB_PASSWORD'),
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        
    cursor = connection.cursor()
    print("Connected to the database.")
    user = UserDetails.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    user_role = user.role
    data = request.get_json()
    print('Received data:', data)
    record = OutputTable.query.filter_by(unique_id=unique_id).first()
    status = data.get('status')
    action = data.get('action')
    romakercomments = data.get('romakercomments')
    rocheckercomments=data.get('rocheckercomments')
    comakercomments=data.get('comakercomments')
    comakerstatus = data.get('comakerstatus')
    comakeraction = data.get('comakeraction')
    cocheckerstatus = data.get('cocheckerstatus')
    cocheckeraction = data.get('cocheckeraction')
    cocheckercomments=data.get('cocheckercomments')
    frm_cell_head =data.get('frm_cell_head')
    
    print("frm_cell_head",frm_cell_head)
    audit_comment= data.get('audit_comment')

    comments=None

    file_objects = []
    if user_role == 'CO Maker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data') 
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'CO Checker' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.co_maker_status = comakerstatus
            
            record.co_maker_action = comakeraction
            record.date_updated_by_co_maker = datetime.now()
            record.comments_by_co_maker = comakercomments
            record.co_maker_files = file_objects
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Co-maker details updated successfully!","redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'CO Checker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'FRM Cell Head' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.co_checker_status=cocheckerstatus
            record.co_checker_action=cocheckeraction
            record.date_updated_by_co_checker=datetime.now()
            record.comments_by_co_checker=cocheckercomments
            record.co_checker_files = file_objects
            db.session.commit()
            return jsonify({"message": "Co-checker details updated successfully!","redirect": "/dashboard"}) 
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'RO Maker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'RO Checker' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.date_updated_by_ro_maker = datetime.now()
            record.comments_by_ro_maker = romakercomments
            record.ro_maker_files = file_objects
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Ro-maker details updated successfully!","redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'RO Checker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'CO Maker' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.comments_by_ro_checker = rocheckercomments
            record.date_updated_by_ro_checker =datetime.now()
            record.ro_checker_files = file_objects
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Ro-checker details updated successfully!", "redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'FRM Cell Head':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = Null , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.frm_cell_head = frm_cell_head
            record.date_updated_by_frm_cell_head =datetime.now()
            record.frm_cell_head_files = file_objects
            print('record1234:', record)
            db.session.commit()
            return jsonify({"message": "FRM Cell Head details updated successfully!", "redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404


    elif user_role == 'Audit':
        print("11111111")
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        
        if record:
            print("222222")
            record.audit_comment = audit_comment
            print("3:",record.audit_comment)
            record.date_updated_by_audit =datetime.now()
            record.audit_files = file_objects
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Audit details updated successfully!", "redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404


@app.route('/Returnedalerts')
@jwt_required()
def Returnedalerts():
    print("inside the function")
    cursor = None
    connection = None
    try:
        user_id = get_jwt_identity()
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        user_role = user.role
        print(user_role,"user")
        print(type(user_role),'usertype')
       

        user_branch_region=user.branch_region
        connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),  
            password=os.getenv('DB_PASSWORD'),
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        
        cursor = connection.cursor()
        print("Connected to the database.")
        cursor.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'output_table' AND column_name = 'distribute';
        """)
        column_check = cursor.fetchone()

        if not column_check:
            cursor.execute("""
                ALTER TABLE output_table 
                ADD COLUMN distribute BOOLEAN DEFAULT FALSE;
            """)
            connection.commit()
            print("Column 'distribute' added to 'output_table' with default value False.")

        if user_role == 'CO Checker':
          query = """
                    SELECT DISTINCT ON (o.unique_id)
                        o.unique_id,  -- Assuming you want distinct results based on the unique_id
                        c.acctno,     
                        c.custcd, 
                        c.customername, 
                        o.branch,
                        o.comments_by_ro_checker,
                        o.ro_maker_files,
                        o.ro_checker_files,
                        o.date_updated_by_ro_checker,
                        o.comments_by_ro_maker,
                        o.date_updated_by_ro_maker,
                        o.query_number AS rule_number,  
                        o.description,    
                        o.*,              
                        o.date AS raised_date,          
                        o.branch_region,
                        u.*, c.*  -- or you can select specific columns from user_details and customers if needed
                    FROM output_table o
                    JOIN user_details u ON o.branch_region = u.branch_region
                    LEFT JOIN customers c ON o.customer_id = c.custcd
                    WHERE o.comments_by_co_maker IS NOT NULL AND o.sendback_to= %s;"""      
          cursor.execute(query, (user_role,)) 
        elif user_role == 'CO Maker':
            query = """
            SELECT DISTINCT ON(o.unique_id)
                o.unique_id,  -- Assuming this is the unique identifier for the alert
                c.acctno,     -- Account number from customers table
                c.custcd,     -- Customer ID from customers table
                c.customername,  -- Customer name from customers table
                o.branch,
                o.query_number AS rule_number,
                o.comments_by_ro_checker,
                o.ro_maker_files,
                o.ro_checker_files,
                o.date_updated_by_ro_checker,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,
                o.description,  -- Description from output table
                o.date AS raised_date,  -- Date (Alert raised date) from output table
                o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.comments_by_ro_checker IS NOT NULL AND o.sendback_to= %s;"""      
            cursor.execute(query, (user_role,))       
        elif user_role == 'RO Checker' and user_branch_region:
            query = f"""
            SELECT DISTINCT ON(o.unique_id)
                o.unique_id,  -- Assuming this is the unique identifier for the alert
                c.acctno,     -- Account number from customers table
                c.custcd,     -- Customer ID from customers table
                c.customername,  -- Customer name from customers table
                o.branch,
                o.query_number AS rule_number,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,
                o.ro_maker_files,
                o.description,  -- Description from output table
                o.date AS raised_date,  -- Date (Alert raised date) from output table
                o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.branch_region = '{user_branch_region}'
            AND o.comments_by_ro_maker IS NOT NULL  
             AND o.sendback_to= %s;"""      
            cursor.execute(query, (user_role,)) 
        elif user_role == 'RO Maker' and user_branch_region:
            query = f"""
            SELECT DISTINCT ON(o.unique_id)
                o.unique_id,  -- Assuming this is the unique identifier for the alert
                c.acctno,     -- Account number from customers table
                c.custcd,     -- Customer ID from customers table
                c.customername,  -- Customer name from customers table
                o.branch,
                
                o.query_number AS rule_number,  -- Rule number (query_no from output table)
                o.description,  -- Description from output table
                o.date AS raised_date,  -- Date (Alert raised date) from output table
                o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd  
            WHERE o.branch_region = '{user_branch_region}'
            AND o.sendback_to= %s AND (o.submited_to='RO Maker' OR o.submited_to IS Null);"""      
            cursor.execute(query, (user_role,))  
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]

        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")


@app.route('/returned_alerts_submit_form/<int:unique_id>', methods=['POST'])
@jwt_required()
def returned_alerts_submit_form(unique_id):     
    print('Received unique_id:', unique_id)
    connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),  
            password=os.getenv('DB_PASSWORD'),                              
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        
    cursor = connection.cursor()
    print("Connected to the database.")
    token = request.headers.get('Authorization').split()[1]
    decoded_token = decode_token(token)
    user_id = get_jwt_identity()
    user = UserDetails.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    user_role = user.role
    data = request.get_json()
    print('Received data:', data)
    record = OutputTable.query.filter_by(unique_id=unique_id).first()
    romakercomments = data.get('romakercomments')
    rocheckercomments=data.get('rocheckercomments')
    comakercomments=data.get('comakercomments')
    comakerstatus = data.get('comakerstatus')
    comakeraction = data.get('comakeraction')
    cocheckerstatus = data.get('cocheckerstatus')
    cocheckeraction = data.get('cocheckeraction')
    cocheckercomments=data.get('cocheckercomments')

    file_objects = []
    if user_role == 'CO Maker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'CO Checker' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")

        if record:
            record.co_maker_status = comakerstatus
            
            record.co_maker_action = comakeraction
            record.date_updated_by_co_maker = datetime.now()
            record.comments_by_co_maker = comakercomments
            record.co_maker_files = file_objects
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Co-maker details updated successfully!","redirect": "/Returnedalerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'CO Checker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'FRM Cell Head' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")

        if record:
            record.co_checker_status=cocheckerstatus
            record.co_checker_action=cocheckeraction
            record.date_updated_by_co_checker=datetime.now()
            record.comments_by_co_checker=cocheckercomments
            record.co_checker_files = file_objects
            db.session.commit()
            return jsonify({"message": "Co-checker details updated successfully!","redirect": "/Returnedalerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'RO Checker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'CO Maker' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")

        if record:
            record.comments_by_ro_checker = rocheckercomments
            record.date_updated_by_ro_checker =datetime.now()
            record.ro_checker_files = file_objects
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Ro-checker details updated successfully!", "redirect": "/Returnedalerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404

    elif user_role == 'RO Maker':
        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        cursor.execute("""
            UPDATE output_table
            SET submited_to = 'RO Checker' , sendback_to=Null
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")

        if record:
            record.date_updated_by_ro_maker = datetime.now()
            record.comments_by_ro_maker = romakercomments
            record.ro_maker_files = file_objects
            record.sendback_to = None
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Ro-maker details updated successfully!","redirect": "/Returnedalerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404



@app.route('/sendback_form/<int:unique_id>', methods=['POST'])
@jwt_required()
def sendback_form(unique_id):
    print('Received unique_id:', unique_id) 
    connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),  
            password=os.getenv('DB_PASSWORD'),
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        
    cursor = connection.cursor()
    print("Connected to the database.")
    token = request.headers.get('Authorization').split()[1]
    decoded_token = decode_token(token)
    user_id = get_jwt_identity()
    user = UserDetails.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    user_role = user.role
    data = request.get_json()
    print('Received dataaa:', data)
    record = OutputTable.query.filter_by(unique_id=unique_id).first()
    status = data.get('status')
    action = data.get('action')
    romakercomments = data.get('romakercomments')
    rocheckercomments=data.get('rocheckercomments')
    comakercomments=data.get('comakercomments')
    comakerstatus = data.get('comakerstatus')
    comakeraction = data.get('comakeraction')
    cocheckerstatus = data.get('cocheckerstatus')
    cocheckeraction = data.get('cocheckeraction')
    cocheckercomments=data.get('cocheckercomments')
    co_checker_sendback_reason=data.get('co_checker_sendback_reason')
    co_maker_sendback_reason=data.get('co_maker_sendback_reason')
    ro_checker_sendback_reason=data.get('ro_checker_sendback_reason')   
    # frm_cell_head =data.get('frm_cell_head')
    frm_cell_head_sendback_reason=data.get('frm_cell_head_sendback_reason')
    print('frmsendbackreason',frm_cell_head_sendback_reason)
    file_objects = []
    if user_role == 'FRM Cell Head':
        print('enteringg')
        
        cursor.execute("""
            UPDATE output_table
            SET sendback_to = 'CO Checker' , submited_to = NULL
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")


        if record:
            record.frm_cell_head_sendback_reason = frm_cell_head_sendback_reason
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "FRM CEll head deatils send back successfully!","redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'CO Checker':
        print('enteringg co checker')
        cursor.execute("""
            UPDATE output_table
            SET sendback_to = 'CO Maker' , submited_to = NULL
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.co_checker_sendback_reason = co_checker_sendback_reason              
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "CO Checker details send back successfully!","redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404                                               
    elif user_role == 'CO Maker':           
        print('enteringg co maker')
        cursor.execute("""
            UPDATE output_table
            SET sendback_to = 'RO Checker' , submited_to = NULL
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.co_maker_sendback_reason = co_maker_sendback_reason
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "CO Maker deatils send back successfully!","redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'RO Checker':
        print('enteringg ro checker')
        cursor.execute("""
            UPDATE output_table
            SET sendback_to = 'RO Maker' , submited_to = NULL
            WHERE unique_id = %s;
        """, (unique_id,))

        # Check if any rows were updated
        if cursor.rowcount == 0:
            print("No rows were updated. Record not found.")
            connection.commit()
            return jsonify({"error": "Record not found!"}), 404

        connection.commit()
        print("Database updated successfully.")
        if record:
            record.ro_checker_sendback_reason = ro_checker_sendback_reason
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "RO Checker deatils send back successfully!","redirect": "/dashboard"})
        else:
            return jsonify({"message": "Record not found!"}), 404



# DB_CONFIG = {
#     "dbname": "FRM",
#     "user": "postgres",
#     "password": "12345",
#     "host": "localhost",
#     "port": "5432",
# }
def distribute_alerts():
    print("entering 2")
    try:
        print("entering 3")
        # Connect to the database
        conn = psycopg2.connect(
            dbname=dbname,
            user=user,
            password=password,
            host=host,
            port=port
        
        )
        # conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()
        # Ensure 'email' and 'distribute' columns exist in 'output_table'
        cursor.execute("""
            DO $$
            BEGIN
                -- Add 'email' column if it doesn't exist
                IF NOT EXISTS (
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_name = 'output_table' AND column_name = 'email'
                ) THEN
                    ALTER TABLE output_table ADD COLUMN email TEXT;
                END IF;
                -- Add 'distribute' column if it doesn't exist
                IF NOT EXISTS (
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_name = 'output_table' AND column_name = 'distribute'
                ) THEN
                    ALTER TABLE output_table ADD COLUMN distribute BOOLEAN DEFAULT FALSE;
                END IF;
            END $$;
        """)
        print("entering 3")
        # Fetch branch_region data from both tables
        cursor.execute("""
            SELECT branch_region, email_id
            FROM user_details
            WHERE role = 'RO Maker'
        """)
        user_data = cursor.fetchall()

        # Create a mapping of branch_region to email_id
        branch_to_email = {row[0]: row[1] for row in user_data}
        print("branch_to_email:", branch_to_email)

        # Fetch data from the output_table
        cursor.execute("SELECT unique_id, branch_region FROM output_table")
        output_data = cursor.fetchall()

        # Update email in output_table where branch_region matches
        for row in output_data:
            print("row:", row)
            print("entering 4")
            record_id, branch_region = row
            email = branch_to_email.get(branch_region)
            if email:
                print("entering 4")
                cursor.execute("""
                    UPDATE output_table
                    SET email = %s, distribute = TRUE
                    WHERE unique_id = %s
                """, (email, record_id))
            else:
                cursor.execute("""
                    UPDATE output_table
                    SET distribute = FALSE
                    WHERE unique_id = %s
                """, (record_id,))

        # Commit changes
        conn.commit()
        return {"status": "success", "message": "Alerts distributed successfully."}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        if conn:
            cursor.close()
            conn.close()
            
@app.route("/distribute_alerts", methods=["POST"])
def handle_distribute_alerts():
    print("entering 1")
    response = distribute_alerts()
    print("entering 6")
    return jsonify(response)



@app.route('/scn_submit_form/<int:unique_id>', methods=['POST'])
@jwt_required()
def scn_submit_form(unique_id):
    print('Received unique_id:', unique_id)
    token = request.headers.get('Authorization').split()[1]
    decoded_token = decode_token(token)
    user_id = get_jwt_identity()
    user = UserDetails.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    user_role = user.role
    data = request.get_json()
    print('Received data:', data)
    record = OutputTable.query.filter_by(unique_id=unique_id).first()
    comakerscncomments=data.get('comakerscncomments')
    comakerscnstatus = data.get('comakerscnstatus')
    cocheckerscnaction = data.get('cocheckerscnaction')
    cocheckerscncomments=data.get('cocheckerscncomments')
    cocheckerscncomments=data.get('cocheckerscncomments')
    frmscncomments= data.get('frm_cell_head_scn_comments')
    frmscnaction=data.get('fraud_cat_report_rbi')

    if user_role == 'CO Maker':
        if record:
            record.co_maker_scn_status = comakerscnstatus
            record.co_maker_scn_date_updated = datetime.now()
            record.co_maker_scn_comments = comakerscncomments
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "Co-maker details updated successfully!","redirect": "/scn-alerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'CO Checker':
        if record:
            record.co_checker_scn_action = cocheckerscnaction
            record.co_checker_scn_date_updated = datetime.now()
            record.co_checker_scn_comments = cocheckerscncomments
            db.session.commit()
            return jsonify({"message": "Co-checker details updated successfully!","redirect": "/scn-alerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    elif user_role == 'FRM Cell Head':
        if record:
            record.fraud_cat_report_rbi = frmscnaction
            record.frm_cell_head_scn_date_updated = datetime.now()
            record.frm_cell_head_scn_comments = frmscncomments
            db.session.commit()
            return jsonify({"message": "FRM Cell head details updated successfully!","redirect": "/scn-alerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404
            


@app.route('/scn_alerts')
@jwt_required()
def scn_alerts():
    try:
        user_id = get_jwt_identity()
        
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        user_role = user.role
        connection = psycopg2.connect(
        dbname=dbname,
        user=os.getenv('DB_USER'),  
        password=os.getenv('DB_PASSWORD'),
        host=os.getenv('DB_HOST'),
        port=os.getenv('DB_PORT')
    )
        
        cursor = connection.cursor()
        
        print("Connected to the database.")
        
        if user_role == 'CO Maker':
            query = """
            SELECT DISTINCT ON (o.unique_id)
                    o.unique_id,  -- Assuming this is the unique identifier for the alert
                    c.acctno,     -- Account number from customers table
                    c.custcd,     -- Customer ID from customers table
                    c.customername,  -- Customer name from customers table
                    o.branch,
                    o.comments_by_ro_checker,
                    o.ro_maker_files,
                    o.ro_checker_files,
                    o.date_updated_by_ro_checker,
                    o.comments_by_ro_maker,
                    o.date_updated_by_ro_maker,
                    o.query_number AS rule_number, 

                    o.comments_by_co_maker,
                    o.date_updated_by_co_maker,
                    o.comments_by_co_checker,
                    o.date_updated_by_co_checker,
                    o.co_maker_files,
                    o.co_checker_files,
                    o.co_maker_status,
                    o.co_maker_action,
                    o.co_checker_status,
                    o.co_checker_action,
                    o.reply_body,
                    o.attachments,
                    o.frm_cell_head,
                    o.date_updated_by_frm_cell_head,
                    o.frm_cell_head_files,
                    o.audit_comment,
                    o.date_updated_by_audit,
                    o.audit_files,
                    o.frm_after_audit,
                    o.frm_after_audit_files,
                    o.frm_after_audit_date,
                    o.description,  -- Description from output table
                    o.date AS raised_date,  -- Date (Alert raised date) from output table
                    o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.frm_cell_head IS NOT NULL 
            AND o.co_maker_scn_comments IS NULL
            AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL);
            """
        elif user_role == 'CO Checker':
            query = """
            SELECT DISTINCT ON (o.unique_id)
                o.unique_id,  -- Assuming you want distinct results based on the unique_id
                c.acctno,     
                c.custcd, 
                c.customername, 
                o.branch,
                o.query_number AS rule_number, 

                o.comments_by_ro_checker,
                o.ro_maker_files,
                o.ro_checker_files,
                o.date_updated_by_ro_checker,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,
                o.query_number AS rule_number, 
                o.frm_cell_head, 
                o.description,    
                o.*,              
                o.date AS raised_date,          
                o.branch_region,
                u.*, c.*  -- or you can select specific columns from user_details and customers if needed
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            
            WHERE o.co_maker_scn_comments IS NOT NULL 
            AND o.co_checker_scn_comments IS NULL
            AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL);
            """
        elif user_role == 'FRM Cell Head':
            query = """
            
SELECT DISTINCT ON (o.unique_id)
                o.unique_id,  -- Assuming you want distinct results based on the unique_id
                c.acctno,     
                c.custcd, 
                c.customername, 
                o.branch,
                o.comments_by_ro_checker,
                o.ro_maker_files,
                o.ro_checker_files,
                o.date_updated_by_ro_checker,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,

                o.query_number AS rule_number,  
                o.description,    
                o.*,              
                o.date AS raised_date,          
                o.branch_region,
                o.co_maker_scn_status,
                o.co_maker_scn_comments,
                o.co_maker_scn_date_updated,
                o.frm_cell_head,
                u.*, c.*  -- or you can select specific columns from user_details and customers if needed
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd 
            WHERE o.co_checker_scn_comments IS NOT NULL 
            AND o.frm_cell_head_scn_comments IS NULL    
            AND o.co_checker_scn_action = 'Account Declared as Fraud Account' -- Fixed the string syntax
            AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL);
            """
        
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]
        # print("Matched Data:", matched_data)  # Debug print
        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")




@app.route('/audit_alerts')
@jwt_required()
def audit_alerts():
    print("111111111111111")
    try:
        print("22222222222")
        user_id = get_jwt_identity()
        
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        user_role = user.role
        connection = psycopg2.connect(
            dbname="FRM",
            user="postgres",
            password="12345",
            host="localhost",
            port="5432"
        )
        cursor = connection.cursor()
        
        print("Connected to the database.")
        
        if user_role == 'FRM Cell Head':
            print("33333333333")
            query = """
            SELECT DISTINCT ON(o.unique_id)
                    o.unique_id,  -- Assuming this is the unique identifier for the alert
                    c.acctno,     -- Account number from customers table
                    c.custcd,     -- Customer ID from customers table
                    c.customername,  -- Customer name from customers table
                    o.branch,
                    o.query_number AS rule_number,
                    o.comments_by_ro_checker,
                    o.ro_maker_files,
                    o.ro_checker_files,
                    o.date_updated_by_ro_checker,
                    o.comments_by_ro_maker,
                    o.date_updated_by_ro_maker,
                    o.comments_by_co_maker,
                    o.date_updated_by_co_maker,
                    o.comments_by_co_checker,
                    o.date_updated_by_co_checker,
                    o.co_maker_files,
                    o.co_checker_files,
                    o.co_maker_status,
                    o.co_maker_action,
                    o.co_checker_status,
                    o.co_checker_action,
                    o.reply_body,
                    o.attachments,
                    o.frm_cell_head,
                    o.date_updated_by_frm_cell_head,
                    o.frm_cell_head_files,
                    o.audit_comment,
                    o.audit_files,
                    o.date_updated_by_audit,
                    o.description,  -- Description from output table
                    o.date AS raised_date,  -- Date (Alert raised date) from output table
                    o.branch_region , -- Branch region from output table,
                    o.email_id
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE (o.audit_comment IS NOT NULL OR o.audit_comment IS NULL)
            AND frm_cell_head IS NOT NULL
            AND frm_after_audit IS  NULL;


            """
        # elif user_role == 'CO Checker':
        #     query = """
        #     SELECT DISTINCT ON (o.unique_id)
        #         o.unique_id,  -- Assuming you want distinct results based on the unique_id
        #         c.acctno,     
        #         c.custcd, 
        #         c.customername, 
        #         o.branch,
        #         o.comments_by_ro_checker,
        #         o.ro_maker_files,
        #         o.ro_checker_files,
        #         o.date_updated_by_ro_checker,
        #         o.comments_by_ro_maker,
        #         o.date_updated_by_ro_maker,
        #         o.query_number AS rule_number, 
        #         o.frm_cell_head, 
        #         o.description,    
        #         o.*,              
        #         o.date AS raised_date,          
        #         o.branch_region,
        #         u., c.  -- or you can select specific columns from user_details and customers if needed
        #     FROM output_table o
        #     JOIN user_details u ON o.branch_region = u.branch_region
        #     LEFT JOIN customers c ON o.customer_id = c.custcd
            
        #     WHERE o.co_maker_scn_comments IS NOT NULL 
        #     AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL);
        #     """
        
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]
        # Debug print: Only show the 'date_updated_by_frm_cell_head' column
        if 'date_updated_by_frm_cell_head' in columns:
            updated_dates = [row.get('date_updated_by_frm_cell_head') for row in matched_data]
            print("Date Updated by FRM Cell Head:", updated_dates)
        else:
            print("The column 'date_updated_by_frm_cell_head' was not found in the query results.")

        # print("Matched Data:", matched_data)  # Debug print
        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")

@app.route('/audit_submit_form/<int:unique_id>', methods=['POST'])
@jwt_required()
def audit_submit_form(unique_id):
    print('Received unique_id:', unique_id)
    token = request.headers.get('Authorization').split()[1]
    decoded_token = decode_token(token)
    user_id = get_jwt_identity()
    user = UserDetails.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    user_role = user.role
    data = request.get_json()
    print('Received data:', data)
    record = OutputTable.query.filter_by(unique_id=unique_id).first()
    frmafteraudit=data.get('frmafteraudit')
    file_objects = []

    if user_role == 'FRM Cell Head':

        for file_data in data.get('files', []):
            filename = file_data.get('filename')
            base64_data = file_data.get('data')
            print('File data:', base64_data)
            file_bytes = base64.b64decode(base64_data)
            base64_encoded_data = base64.b64encode(file_bytes).decode('utf-8')
            file_objects.append({
                "filename": filename,
                "filedata": base64_encoded_data,
            })
            print('file_objects:', file_objects)
        if record:
            record.frm_after_audit = frmafteraudit
            record.frm_after_audit_files=file_objects
            record.frm_after_audit_date=datetime.now()
            print('record:', record)
            db.session.commit()
            return jsonify({"message": "frm after audit details updated successfully!","redirect": "/audit_alerts"})
        else:
            return jsonify({"message": "Record not found!"}), 404
    # elif user_role == 'CO Checker':
    #     if record:
    #         record.co_checker_scn_action = cocheckerscnaction
    #         record.co_checker_scn_date_updated = datetime.now()
    #         record.co_checker_scn_comments = cocheckerscncomments
    #         db.session.commit()
    #         return jsonify({"message": "Co-checker details updated successfully!","redirect": "/dashboard"})
    #     else:
    #         return jsonify({"message": "Record not found!"}), 404






# email reply checking part starting    
# stop_thread = False
# monitor_thread = None
# def check_email_replies():
#     try:
#         mail = imaplib.IMAP4_SSL(app.config['IMAP_SERVER'], app.config['IMAP_PORT'])
#         mail.login(app.config['EMAIL_ACCOUNT'], app.config['EMAIL_PASSWORD'])
#         mail.select("inbox")
#         subject_keyword = 'Examination of irregularities observed in the Loan A/c'
#         status, messages = mail.search(None, f'SUBJECT "{subject_keyword}"')
#         email_ids = messages[0].split()
#         if not email_ids:
#             print("No new emails found.")
#             mail.logout()
#             return
#         print("New emails:", email_ids)
#         for email_id in email_ids:
#             status, msg_data = mail.fetch(email_id, '(RFC822)')
#             for response_part in msg_data:
#                 if isinstance(response_part, tuple):
#                     msg = email.message_from_bytes(response_part[1])
#                     subject, encoding = decode_header(msg["Subject"])[0]
#                     if isinstance(subject, bytes):
#                         subject = subject.decode(encoding or 'utf-8', errors='ignore')
#                     reply_body = ""
#                     # attachments = []
#                     file_objects = []
#                     if msg.is_multipart():
#                         for part in msg.walk():
#                             content_type = part.get_content_type()
#                             content_disposition = str(part.get("Content-Disposition"))
#                             if content_type == "text/plain" and "attachment" not in content_disposition:
#                                 reply_body = part.get_payload(decode=True).decode(errors='ignore')
#                             if "attachment" in content_disposition:
#                                 filename = part.get_filename()
#                                 print(filename,"====================")
#                                 if filename:
#                                     # added
#                                     try:
#                                         filename = decode_header(filename)[0][0]
#                                         if isinstance(filename, bytes):
#                                             filename = filename.decode('utf-8', errors='ignore')
#                                     except Exception as e:
#                                         print(f"Error decoding filename: {e}")
#                                         filename = "unknown_filename"
#                                     filename = secure_filename(filename)
#                                     ####

#                                     # # filename = secure_filename(filename)
#                                     # print("file name is:",filename)
#                                     # file_data = part.get_payload(decode=True)
#                                     # attachments.append({
#                                     #     "filename": filename,
#                                     #     "file_data": file_data
#                                     # })

#                                     file_data = part.get_payload(decode=True)
#                                     base64_encoded_data = base64.b64encode(file_data).decode('utf-8')
#                                     file_objects.append({
#                                         "filename": filename,
#                                         "filedata": base64_encoded_data,
#                                     })
#                                     print(f"Processed attachment: {filename}")
#                     else:
#                         reply_body = msg.get_payload(decode=True).decode(errors='ignore')


#                     clean_reply_body = reply_body
#                     clean_reply_body = re.split(r"\nOn .* at .* wrote:\n", clean_reply_body)[0]
#                     clean_reply_body = re.split(r"\nOn .* wrote:", clean_reply_body)[0]
#                     clean_reply_body = re.split(r"\nFrom:.*\n", clean_reply_body)[0]
#                     clean_reply_body = re.split(r"\nSent:.*\n", clean_reply_body)[0]
#                     clean_reply_body = re.split(r"\nTo:.*\n", clean_reply_body)[0]
#                     clean_reply_body = re.split(r"\nSubject:.*\n", clean_reply_body)[0]
#                     clean_reply_body = "\n".join(
#                         line for line in clean_reply_body.splitlines() if not line.startswith(">")
#                     ).strip()
#                     clean_reply_body = re.sub(r"^\s*[-]{2,}.*$", "", clean_reply_body, flags=re.MULTILINE)
#                     clean_reply_body = clean_reply_body.strip()
#                     unique_id = extract_unique_id_from_subject(subject)
#                     print(f"Unique ID from subject: {unique_id}")
#                     if unique_id:
#                         with app.app_context():
#                             customer_data = OutputTable.query.filter_by(unique_id=unique_id).first()
#                             if customer_data:
#                                 if not customer_data.reply_body:
#                                     customer_data.reply_body = clean_reply_body
#                                 # if attachments:
#                                 #     customer_data.attachments = [{"filename": attachment["filename"], "file_data": attachment["file_data"].decode()} for attachment in attachments]
#                                 #     db.session.commit()
#                                 #     print(f"Stored reply for unique_id {unique_id}")
#                                 # else:
#                                 #     print(f"Reply already stored for unique_id {unique_id}")
#                                 if file_objects:
#                                     customer_data.attachments = file_objects
                                    
#                                     print(f"Stored reply and attachments for unique_id {unique_id}")
#                                 else:
#                                     print(f"No attachments to store for unique_id {unique_id}")
#                                 db.session.commit()

#                             else:
#                                 print(f"No matching record found for unique_id {unique_id}")
#         mail.logout()
#     except Exception as e:
#         print(f"Error checking email replies: {e}")


# def extract_unique_id_from_subject(subject):
#     try:
#         if "ID: " in subject:
#             unique_id = subject.split("ID: ")[1]
#             return unique_id
#     except IndexError:
#         print("Unique ID format in subject line is incorrect.")
#     return None
# def monitor_inbox():
#     while not stop_thread:  
#         check_email_replies()
#         time.sleep(30)  
# def start_monitoring():
#     global monitor_thread
#     monitor_thread = threading.Thread(target=monitor_inbox)
#     monitor_thread.daemon = True
#     monitor_thread.start()
# stop_attempted = False
# def stop_monitoring():
#     global stop_thread, stop_attempted
#     if not stop_attempted:
#         stop_thread = True
#         monitor_thread.join()  
#         stop_attempted = True
#         print("Monitoring thread stopped.")
#     else:
#         print("Thread has already been stopped.")
# atexit.register(stop_monitoring)
# start_monitoring()

# from werkzeug.utils import secure_filename
# email reply checking part ending

@app.route('/reports')
@jwt_required()
def reports():
    try:
        user_id = get_jwt_identity()
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        user_role = user.role
        connection = psycopg2.connect(
    dbname=dbname,
    user=os.getenv('DB_USER'),  # Correct
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT')
)
        cursor = connection.cursor()
        
        print("Connected to the database.")
        
        if user_role == 'FRM Cell Head':
            query = """
            SELECT DISTINCT ON (o.unique_id)
                o.unique_id,                     -- Unique identifier for the alert
                c.acctno,                        -- Account number from customers table
                c.custcd,                        -- Customer ID from customers table
                c.customername,                  -- Customer name from customers table
                o.branch,                        -- Branch details
                o.comments_by_ro_checker,        -- Comments by RO Checker
                o.query_number AS rule_number,  

                o.ro_maker_files,                -- RO Maker Files
                o.ro_checker_files,              -- RO Checker Files
                o.date_updated_by_ro_checker,    -- Date updated by RO Checker
                o.comments_by_ro_maker,          -- Comments by RO Maker
                o.date_updated_by_ro_maker,      -- Date updated by RO Maker
                o.comments_by_co_maker,          -- Comments by CO Maker
                o.date_updated_by_co_maker,      -- Date updated by CO Maker
                o.comments_by_co_checker,        -- Comments by CO Checker
                o.date_updated_by_co_checker,    -- Date updated by CO Checker
                o.co_maker_files,                -- CO Maker Files
                o.co_checker_files,              -- CO Checker Files
                o.co_maker_status,               -- CO Maker Status
                o.co_maker_action,               -- CO Maker Action
                o.co_checker_status,             -- CO Checker Status
                o.co_checker_action,             -- CO Checker Action
                o.reply_body,                    -- Reply body
                o.lc_issued,               -- Attachments
                o.frm_cell_head,                 -- Form Cell Head
                o.description,                   -- Description from output table
                o.date AS raised_date,           -- Alert raised date
                o.branch_region,                 -- Branch region
                o.branch_code,                   -- Branch code
                o.lc_beneficiary_bank_name,      -- LC Beneficiary Bank Name
                o.pan_number,                    -- PAN Number
                u.username,                      -- Username from user details
                u.role,                          -- Role from user details
                o.frm_cell_head_scn_date_updated,
                o.query_number,
                c.accttype,
                o.branch_region,

                -- Newly added columns from output_table
                o.account_number,
                o.issue_party_code,
                o.lc_beneficiary_party,
                o.lc_beneficiary_bank_code,
                o.lc_beneficiary_acc_no,
                o.lc_issue_type,
                o.lc_ref_no,
                o.lc_issue_date,
                o.lc_expiry_date,
                o.trade_transaction_beneficiary_party,
                o.trade_transaction_bank_name,
                o.trade_transaction_beneficiary_acc_no,
                o.aadharno,
                o.email_id,
                o.customer_id,
                o.date,
                o.co_maker_status,
                o.co_maker_action,
                o.co_maker_scn_status,
                o.co_maker_scn_comments,
                o.co_maker_scn_date_updated,
                o.co_checker_status,
                o.co_checker_action,
                o.co_checker_scn_action,
                o.co_checker_scn_comments,
                o.co_checker_scn_date_updated,
                o.reply_body,
                o.ro_maker_files,
                o.ro_checker_files,
                o.co_maker_files,
                o.co_checker_files,
                o.attachments,
                o.distribute,
                o.date_updated_by_frm_cell_head,
                o.audit_comment,
                o.date_updated_by_audit,
                o.frm_cell_head_files,
                o.audit_files,
                o.frm_cell_head_scn_comments,
                o.fraud_cat_report_rbi,
                o.frm_after_audit,
                o.frm_after_audit_files,
                o.frm_after_audit_date,
                o.adhoc_loan_request_id,
                o.remarks,
                o.currency_code,
                o.adhoc_date_of_request,
                o.adhoc_date_of_sanction,
                o.purpose,
                o.attachments, 
                o.loan_approval_status

                                        -- Include all other columns from output_table
            FROM 
                output_table o
            JOIN 
                user_details u 
                ON o.branch_region = u.branch_region
            LEFT JOIN 
                customers c 
                ON o.customer_id = c.custcd
           WHERE 
                (
                    (
                        o.frm_cell_head_scn_comments IS NOT NULL        
                        AND o.fraud_cat_report_rbi IS NOT NULL  
                    )
                    OR
                    (
                        (o.audit_comment IS NOT NULL OR o.audit_comment IS NULL)
                        AND frm_cell_head IS NOT NULL
                        AND frm_after_audit IS NULL
                    )
                );


            

            """
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        # print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]
        print("Matched Data:", matched_data)  # Debug print
        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")


import zipfile

DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'Downloads') 
app.config["DOWNLOAD_FOLDER"] = DOWNLOAD_FOLDER


@app.route('/cocheckerreturned')
@jwt_required()
def cocheckerreturned():
    try:
        user_id = get_jwt_identity()
        
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        user_role = user.role
        connection = psycopg2.connect(
    dbname=dbname,
    user=os.getenv('DB_USER'),  # Correct
    password=os.getenv('DB_PASSWORD'),
    host=os.getenv('DB_HOST'),
    port=os.getenv('DB_PORT')
)
        cursor = connection.cursor()
        
        print("Connected to the database.")
        
        if user_role == 'CO Maker':
            query = """
            SELECT DISTINCT ON(o.unique_id)
                    o.unique_id,  -- Assuming this is the unique identifier for the alert
                    c.acctno,     -- Account number from customers table
                    c.custcd,     -- Customer ID from customers table
                    c.customername,  -- Customer name from customers table
                    o.branch,
                    o.comments_by_ro_checker,
                    o.ro_maker_files,
                    o.ro_checker_files,
                    o.date_updated_by_ro_checker,
                    o.query_number AS rule_number,  -- Rule number (query_no from output table)
                    o.comments_by_ro_maker,
                    o.date_updated_by_ro_maker,
                    o.comments_by_co_maker,
                    o.date_updated_by_co_maker,
                    o.comments_by_co_checker,
                    o.date_updated_by_co_checker,
                    o.co_maker_files,
                    o.co_checker_files,
                    o.co_maker_status,
                    o.co_maker_action,
                    o.co_checker_status,
                    o.co_checker_action,
                    o.reply_body,
                    o.attachments,
                    o.frm_cell_head,
                    o.description,  -- Description from output table
                    o.date AS raised_date,  -- Date (Alert raised date) from output table
                    o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.co_checker_status IN ('Reject');
            """
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]
        # print("Matched Data:", matched_data)  # Debug print
        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")



DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'Downloads')
app.config["DOWNLOAD_FOLDER"] = DOWNLOAD_FOLDER

@app.route('/download-report/<int:unique_id>', methods=['GET', 'POST'])
@jwt_required()
def download_report(unique_id):
    try:
        print('Received unique_id:', unique_id)             
        auth_header = request.headers.get('Authorization')
        if not auth_header or len(auth_header.split()) < 2:
            return jsonify({"error": "Authorization header missing or invalid"}), 401
        token = auth_header.split()[1]
        decoded_token = decode_token(token)
        user_id = get_jwt_identity()
        user = db.session.get(UserDetails, user_id)

        if not user:
            return jsonify({"error": "User not found"}), 404
        user_role = user.role
        
        data = request.get_json()
        print('customerid:',data.get('customerId'))
        # print('dataaaaa:',data)
        branch = data.get("branch")
        branch_code = data.get("branch_code")
        bank_name = data.get("bank_name")
        cust_name = data.get("customername")
        pan = data.get("pan_number")
        frm_scn_date = data.get('frm_cell_head_scn_date_updated')       
        branch_region=data.get('branch_region') 
        source=data.get('source')
        print('branch_region:',branch_region)
        print('frm_scn_date:', frm_scn_date)
        print('source:',source)
        romakerfiles = data.get('ro_maker_files')
        rocheckerfiles = data.get('ro_checker_files')
        comakerfiles = data.get('co_maker_files')
        cocheckerfiles = data.get('co_checker_files')
        attachments = data.get('attachments')
        frm_cell_head_files = data.get('frm_cell_head_files')
        audit_files = data.get('audit_files')
        frm_after_audit_files = data.get('frm_after_audit_files')

        # Database connection setup
        connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),  # Correct
            password=os.getenv('DB_PASSWORD'),
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        cursor = connection.cursor()

        print("Connected to the database.")

        query = """ 
            SELECT *
            FROM 
                user_details
            WHERE 
                branch_region = %s;
            """

        cursor.execute(query, (branch_region,))
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        matched_data = [dict(zip(columns, row)) for row in results]
        user_roles = [{"username": row["username"], "role": row["role"]} for row in matched_data]
        query1 = """ 
            SELECT *
            FROM 
                user_details
            WHERE 
                role NOT IN ('RO Maker', 'RO Checker');
            """
        
        # Execute the query with parameter
        cursor.execute(query1)
        results1 = cursor.fetchall()
        columnss = [desc[0] for desc in cursor.description]
        # print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columnss, row)) for row in results1]
        roles = [{"username": row["username"], "role": row["role"]} for row in matched_data]

        # Print the extracted data for verification
        print("Extracted Usernames and Roles:", roles)  
        # print(matched_data,':matcheddata')

        if source == 'CRILIC':
            print('Processing CRILIC source.')

            wb = Workbook()
            ws = wb.active
            ws.title = "User Comments and Output Details"

            # Add headers for the Excel file
            ws.append([
                "Unique Id", "Customer Id", "Account Number", "Issue Party Code", "lc Issued", "lc Beneficiary Party",
                "lc Beneficiary bank Name", "lc Beneficiary bank Code", "lc Beneficiary bank Acctno", "lc Issue Type",
                "lc Ref No", "lc Issue Date", "lc Expiry Date", "Trade Transaction Benificiary Party", "Trade Transaction Bank Name",
                "Trade Transaction Benificiary AcctNo", "Customer Name", "PAN Number", "Adhaar No", "Branch", "Branch Region",
                "Branch Code", "Email Id", "Date", "Query number", "Description", "Comments By Ro Maker", "Ro Maker Submited date",
                "Comments By Ro Checker", "Ro Checker Submited date", "Comments By Co Maker", "Co Maker Submited date", "Co Maker Status",
                "Co Maker Action", "Co Maker SCN Status", "Co Maker SCN Comments", "Co maker SCN Submited Date", "Comments By Co Checker",
                "Co Checker Submited date", "Co Checker Status", "Co Checker Action", "Co Checker SCN Action", "Co Checker SCN Comments",
                "Co Checker SCN Submited Date", "Reply Body", "Distribute", "Frm Cell Head", "Date Updated by FRM Cell Head",
                "Audit Comment", "Audit Submited Date", "FRM Cell head SCN Comments", "FRM Cell Head SCN submited date",
                "Fraud Cat Report RBI", "FRM After Audit Comments", "FRM After Audit Date", "Adhoc Loan Request Id", "Remarks",
                "Currency Code", "Adhoc Date Of Request", "Adhoc Date Of Sanction", "Purpose", "Loan Approve Status", "Username",
                "Role", "Bank Name", "Raised Date"
            ])

            # for user in roles:
            ws.append([
                unique_id, data.get('customerId'), data.get('account_number'), data.get('issue_party_code'), data.get('lc_issued'),
                data.get('lc_beneficiary_party'), data.get('lc_beneficiary_bank_name'), data.get('lc_beneficiary_bank_code'),
                data.get('lc_beneficiary_acc_no'), data.get('lc_issue_type'), data.get('lc_ref_no'), data.get('lc_issue_date'),
                data.get('lc_expiry_date'), data.get('trade_transaction_beneficiary_party'), data.get('trade_transaction_bank_name'),
                data.get('trade_transaction_beneficiary_acc_no'), data.get("customername"), data.get("pan_number"), data.get('aadharno'),
                data.get("branch"), data.get('branch_region'), data.get("branch_code"), data.get('email_id'), data.get('date'),
                data.get('query_number'), data.get("description"), data.get('comments_by_ro_maker'), data.get('date_updated_by_ro_maker'),
                data.get('comments_by_ro_checker'), data.get('date_updated_by_ro_checker'), data.get('comments_by_co_maker'),
                data.get('date_updated_by_co_maker'), data.get('co_maker_status'), data.get('co_maker_action'), data.get('co_maker_scn_status'),
                data.get('co_maker_scn_comments'), data.get('co_maker_scn_date_updated'), data.get('comments_by_co_checker'),
                data.get('date_updated_by_co_checker'), data.get('co_checker_status'), data.get('co_checker_action'),
                data.get('co_checker_scn_action'), data.get('co_checker_scn_comments'), data.get('co_checker_scn_date_updated'),
                data.get('reply_body'), data.get('distribute'), data.get('frm_cell_head'), data.get('date_updated_by_frm_cell_head'),
                data.get('audit_comment'), data.get('date_updated_by_audit'), data.get('frm_cell_head_scn_comments'),
                data.get('frm_cell_head_scn_date_updated'), data.get('fraud_cat_report_rbi'), data.get('frm_after_audit'),
                data.get('frm_after_audit_date'), data.get('adhoc_loan_request_id'), data.get('remarks'), data.get('currency_code'),
                data.get('adhoc_date_of_request'), data.get('adhoc_date_of_sanction'), data.get('purpose'), data.get('loan_approval_status'),
                    data.get("bank_name"), data.get("raised_date")
            ])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Prepare additional files for inclusion in the ZIP
            file_data_map = {
                "attachments":attachments,
                "romakerfiles": romakerfiles,
                "rocheckerfiles": rocheckerfiles,
                "comakerfiles": comakerfiles,
                "cocheckerfiles": cocheckerfiles,
                "frm_cell_head_files": frm_cell_head_files,
                "audit_files": audit_files,
                "frm_after_audit_files": frm_after_audit_files,
            }

            file_buffers = {}
            for file_key, file_items in file_data_map.items():
                if file_items:
                    for file in file_items:
                        filename = file.get("filename")
                        filedata = file.get("filedata")
                        if filename and filedata:
                            decoded_data = filedata if isinstance(filedata, bytes) else base64.b64decode(filedata)
                            file_buffers[filename] = BytesIO(decoded_data)

            # Generate the ZIP file
            # Generate the ZIP file
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                # Add the Excel file to the ZIP
                zip_file.writestr("user_comments_output.xlsx", output.getvalue())
                
                # Add files to specific folders
                for file_key, file_items in file_data_map.items():
                    if file_items:
                        for file in file_items:
                            filename = file.get("filename")
                            filedata = file.get("filedata")
                            if filename and filedata:
                                # Decode the file data
                                decoded_data = filedata if isinstance(filedata, bytes) else base64.b64decode(filedata)
                                # Write the file into its respective folder
                                folder_name = file_key  # Use the key as the folder name
                                zip_file.writestr(f"{folder_name}/{filename}", decoded_data)

            zip_buffer.seek(0)

            # Return the ZIP file
            return send_file(
                zip_buffer,
                as_attachment=True,
                download_name="output_files.zip",
                mimetype="application/zip"
            )

        elif source=='Reserve Bank of India':
            file_name = "ReserveBankOfIndia.XLSX"
            file_path = os.path.join(app.config["DOWNLOAD_FOLDER"], file_name)

            if not os.path.exists(file_path):
                app.logger.error(f"File not found at {file_path}")
                return jsonify({"error": "File not found. Please upload the required file."}), 404

            # Load workbook
            wb = load_workbook(file_path)
            print('wbb:', wb)

            # Insert data into specific sheets
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print('sheet:', sheet)
                for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=8):  # Adjust row/col range as needed
                    for cell in row:
                        print(f"Cell {cell.coordinate}: {cell.value}")

                if sheet_name == "FMR III (FUA) Sheet 1":
                    sheet.cell(row=6, column=3, value=data.get("bank_name"))  # C2
                    sheet.cell(row=7, column=3, value=data.get("query_number"))
                    # sheet.cell(row=8, column=3, value=data.get("branch")) 
                    sheet.cell(row=9, column=3, value=data.get("branch_code")) 
                    sheet.cell(row=10, column=3, value=data.get("branch"))
                    sheet.cell(row=12, column=3, value=data.get("accttype")) 
                    sheet.cell(row=14, column=3, value=data.get("pan_number"))
                    sheet.cell(row=25, column=3, value=data.get("description"))
                    sheet.cell(row=31, column=3, value=data.get("raised_date"))    
                    # C3 
                # if sheet_name == "FMR III (FUA) Sheet 2":
                #     current_row = 7  # Starting row for RO Maker

                #     for user in user_roles + roles:  # Combine both role lists for processing
                #         role = user['role']
                #         # Check if the current row is already occupied, and if so, insert two new rows
                #         if any(sheet.cell(row=current_row, column=col).value for col in range(1, 14)):
                #             sheet.insert_rows(current_row)  # Insert a blank row
                #             sheet.insert_rows(current_row)  # Insert another blank row
                #             current_row += 2  # Adjust current row to skip the newly inserted rows

                #         # Fill in role details
                #         if role == 'RO Maker':
                #             sheet.cell(row=current_row, column=2, value=1)                      
                #             sheet.cell(row=current_row, column=3, value=user.get("username"))
                #             sheet.cell(row=current_row, column=5, value=role)
                #             sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                #             sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_ro_maker"))
                #             sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                #             sheet.cell(row=current_row, column=13, value=data.get("comments_by_ro_maker"))

                #         elif role == 'RO Checker':  # Ensure RO Checker details exist
                #             sheet.cell(row=current_row, column=2, value=2)
                #             sheet.cell(row=current_row, column=3, value=user.get("username"))
                #             sheet.cell(row=current_row, column=5, value=role)
                #             sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                #             sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_ro_checker"))
                #             sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                #             sheet.cell(row=current_row, column=13, value=data.get("comments_by_ro_checker"))

                #         elif role == 'CO Maker':
                #             sheet.cell(row=current_row, column=2, value=3)
                #             sheet.cell(row=current_row, column=3, value=user.get("username"))
                #             sheet.cell(row=current_row, column=5, value=role)
                #             sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                #             sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_co_maker"))
                #             sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                #             sheet.cell(row=current_row, column=13, value=data.get("comments_by_co_maker"))

                #         elif role == 'CO Checker':
                #             sheet.cell(row=current_row, column=2, value=4)
                #             sheet.cell(row=current_row, column=3, value=user.get("username"))
                #             sheet.cell(row=current_row, column=5, value=role)
                #             sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                #             sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_co_checker"))
                #             sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                #             sheet.cell(row=current_row, column=13, value=data.get("comments_by_co_checker"))

                #         elif role == 'FRM Cell Head':
                #             sheet.cell(row=current_row, column=2, value=5)
                #             sheet.cell(row=current_row, column=3, value=user.get("username"))
                #             sheet.cell(row=current_row, column=5, value=role)
                #             sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                #             sheet.cell(row=current_row, column=10, value='-')
                #             sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                #             sheet.cell(row=current_row, column=13, value=data.get("frmcellhead"))

                #         # Move to the next alternative row for the next role
                #         current_row += 2


                if sheet_name == "FMR III (FUA) Sheet 2":
                    current_row = 6 # Starting row for data entry
                
                    for user in user_roles:
                        role = user['role']
                        if role == 'RO Maker':
                            sheet.cell(row=current_row, column=2, value=1)
                            sheet.cell(row=current_row, column=3, value=user.get("username"))
                            sheet.cell(row=current_row, column=5, value=role)
                            sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                            sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_ro_maker"))
                            sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                            sheet.cell(row=current_row, column=13, value=data.get("comments_by_ro_maker"))

                        elif role == 'RO Checker':
                            sheet.cell(row=current_row, column=2, value=2)
                            sheet.cell(row=current_row, column=3, value=user.get("username"))
                            sheet.cell(row=current_row, column=5, value=role)
                            sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                            sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_ro_checker"))
                            sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                            sheet.cell(row=current_row, column=13, value=data.get("comments_by_ro_checker"))

                        current_row += 1

                    for user in roles:
                        role = user['role']
                        # Check if the next row is already occupied and insert a new row if necessary
                        if any(sheet.cell(row=current_row, column=col).value for col in range(1, 14)):  # Check if any cell in the row is occupied
                            sheet.insert_rows(current_row)  # Insert a blank row

                        if role == 'CO Maker':
                            sheet.cell(row=current_row, column=2, value=3)
                            sheet.cell(row=current_row, column=3, value=user.get("username"))
                            sheet.cell(row=current_row, column=5, value=role)
                            sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                            sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_co_maker"))
                            sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                            sheet.cell(row=current_row, column=13, value=data.get("comments_by_co_maker"))

                        elif role == 'CO Checker':
                            sheet.cell(row=current_row, column=2, value=4)
                            sheet.cell(row=current_row, column=3, value=user.get("username"))
                            sheet.cell(row=current_row, column=5, value=role)
                            sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                            sheet.cell(row=current_row, column=10, value=data.get("date_updated_by_co_checker"))
                            sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                            sheet.cell(row=current_row, column=13, value=data.get("comments_by_co_checker"))

                        elif role == 'FRM Cell Head':
                            sheet.cell(row=current_row, column=2, value=5)
                            sheet.cell(row=current_row, column=3, value=user.get("username"))
                            sheet.cell(row=current_row, column=5, value=role)
                            sheet.cell(row=current_row, column=7, value=data.get("raised_date"))
                            sheet.cell(row=current_row, column=10, value='-')
                            sheet.cell(row=current_row, column=11, value=data.get("frm_cell_head_scn_date_updated"))
                            sheet.cell(row=current_row, column=13, value=data.get("frm_cell_head"))

                        current_row += 1    

                
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"Processing sheet: {sheet_name}")
                merged_cells = sheet.merged_cells
                for merged_range in sheet.merged_cells.ranges:
                    print(f"Processing merged range: {merged_range}")
                    
                    # Extract the boundaries of the merged range
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    print(f"Boundaries of merged range: {min_col}, {min_row}, {max_col}, {max_row}")
                    
                    # Modify only the top-left cell of the merged range
                    top_left_cell = sheet.cell(row=min_row, column=min_col)
                    print(f"Modifying top-left cell at {top_left_cell.coordinate}, Current Value: {top_left_cell.value}")
                    # top_left_cell.value = "New Value"  # Replace this with your desired value

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Return the updated file
            return send_file(
                output,
                as_attachment=True,
                download_name="formatted_report.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        app.logger.error(f"An error occurred: {e}")
        app.logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500



@app.route('/submitted_alerts')
@jwt_required()
def submitted_alerts():
    print("it is entering to submitted alerts ")
    cursor = None
    connection = None
    try:
        user_id = get_jwt_identity()
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        user_role = user.role
        user_branch_region = user.branch_region
        
        connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),  
            password=os.getenv('DB_PASSWORD'),
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        
        cursor = connection.cursor()
        print("Connected to the database.")

        # Role-based query generation
        role_query_map = {
            'RO Maker': "o.comments_by_ro_maker IS NOT NULL",
            'RO Checker': "o.comments_by_ro_checker IS NOT NULL",
            'CO Maker': "o.comments_by_co_maker IS NOT NULL",
            'CO Checker': "o.comments_by_co_checker IS NOT NULL AND o.co_checker_status IN ('Recommended For making as RFA', 'Under Investigation');",
            'FRM Cell Head': "o.frm_cell_head IS NOT NULL",
            'Audit': "o.audit_comment IS NOT NULL"
        }

        if user_role not in role_query_map:
            return jsonify({"error": "Invalid role or role not supported."}), 400

        comments_condition = role_query_map[user_role]

        # Build query with specific conditions
        query = f"""
        SELECT DISTINCT ON (o.unique_id)
            o.unique_id, 
            o.query_number AS rule_number,
            o.description,
            o.date AS raised_date,
            o.branch_region,
            c.acctno,
            c.custcd,
            c.customername,
            o.*
        FROM output_table o
        JOIN user_details u ON o.branch_region = u.branch_region
        LEFT JOIN customers c ON o.customer_id = c.custcd
        WHERE {comments_condition}
        """
        
        # Additional branch region filtering for roles with branch-specific permissions
        if user_role in ['RO Maker', 'RO Checker', 'CO Maker', 'CO Checker'] and user_branch_region:
            query += f" AND o.branch_region = '{user_branch_region}'"

        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        print(f"Fetched {len(results)} rows from the database.")

        matched_data = [dict(zip(columns, row)) for row in results]
        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed.")



@app.route('/scn_submitted_alerts')
@jwt_required()
def scn_submitted_alerts():
    try:
        user_id = get_jwt_identity()
        
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        user_role = user.role
        connection = psycopg2.connect(
        dbname=dbname,
        user=os.getenv('DB_USER'),  
        password=os.getenv('DB_PASSWORD'),
        host=os.getenv('DB_HOST'),
        port=os.getenv('DB_PORT')
    )
        
        cursor = connection.cursor()
        
        print("Connected to the database.")
        
        if user_role == 'CO Maker':
            query = """
            SELECT DISTINCT
                    o.unique_id,  -- Assuming this is the unique identifier for the alert
                    c.acctno,     -- Account number from customers table
                    c.custcd,     -- Customer ID from customers table
                    c.customername,  -- Customer name from customers table
                    o.branch,
                    o.comments_by_ro_checker,
                    o.ro_maker_files,
                    o.ro_checker_files,
                    o.date_updated_by_ro_checker,
                    o.comments_by_ro_maker,
                    o.date_updated_by_ro_maker,
                    o.query_number AS rule_number, 

                    o.comments_by_co_maker,
                    o.date_updated_by_co_maker,
                    o.comments_by_co_checker,
                    o.date_updated_by_co_checker,
                    o.co_maker_files,
                    o.co_checker_files,
                    o.co_maker_status,
                    o.co_maker_action,
                    o.co_checker_status,
                    o.co_checker_action,
                    o.reply_body,
                    o.attachments,
                    o.frm_cell_head,
                    o.date_updated_by_frm_cell_head,
                    o.frm_cell_head_files,
                    o.audit_comment,
                    o.date_updated_by_audit,
                    o.audit_files,
                    o.frm_after_audit,
                    o.frm_after_audit_files,
                    o.frm_after_audit_date,
                    o.reply_body,
                    o.attachments,
                    o.co_maker_scn_comments,
                    o.co_maker_scn_status,
                    o.co_maker_scn_date_updated,
                    o.description,  -- Description from output table
                    o.date AS raised_date,  -- Date (Alert raised date) from output table
                    o.branch_region  -- Branch region from output table
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.frm_cell_head IS NOT NULL 
            AND o.co_maker_scn_comments IS NOT NULL
            AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL);
            """
        elif user_role == 'CO Checker':
            query = """
            SELECT DISTINCT ON (o.unique_id)
                o.unique_id,  -- Assuming you want distinct results based on the unique_id
                c.acctno,     
                c.custcd, 
                c.customername, 
                o.branch,
                o.query_number AS rule_number, 

                o.comments_by_ro_checker,
                o.ro_maker_files,
                o.ro_checker_files,
                o.date_updated_by_ro_checker,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,
                o.query_number AS rule_number, 
                o.frm_cell_head, 
                o.description,    
                o.*,              
                o.date AS raised_date,          
                o.branch_region,
                u.*, c.*  -- or you can select specific columns from user_details and customers if needed
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE o.co_maker_scn_comments IS NOT NULL 
            AND o.co_checker_scn_comments IS NOT NULL
            AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL)
            AND o.co_checker_scn_action = 'Account Declared as Fraud Account';
            """
        elif user_role == 'FRM Cell Head':
            query = """
            
SELECT DISTINCT ON (o.unique_id)
                o.unique_id,  -- Assuming you want distinct results based on the unique_id
                c.acctno,     
                c.custcd, 
                c.customername, 
                o.branch,
                o.comments_by_ro_checker,
                o.ro_maker_files,
                o.ro_checker_files,
                o.date_updated_by_ro_checker,
                o.comments_by_ro_maker,
                o.date_updated_by_ro_maker,

                o.query_number AS rule_number,  
                o.description,    
                o.*,              
                o.date AS raised_date,          
                o.branch_region,
                o.co_maker_scn_status,
                o.co_maker_scn_comments,
                o.co_maker_scn_date_updated,
                o.frm_cell_head,
                u.*, c.*  -- or you can select specific columns from user_details and customers if needed
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd 
            WHERE o.co_checker_scn_comments IS NOT NULL 
            AND o.frm_cell_head_scn_comments IS NOT NULL    
            AND o.co_checker_scn_action = 'Account Declared as Fraud Account' -- Fixed the string syntax
            AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL);
            """
        
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]
        # print("Matched Data:", matched_data)  # Debug print
        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")


@app.route('/submitted_alerts_frm_after_audit')
@jwt_required()
def submitted_alerts_frm_after_audit():
    print("111111111111111")
    try:
        print("22222222222")
        user_id = get_jwt_identity()
        
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        user_role = user.role
        connection = psycopg2.connect(
            dbname="FRM",
            user="postgres",
            password="12345",  
            host="localhost",
            port="5432"
        )
        cursor = connection.cursor()
        
        print("Connected to the database.")
        
        if user_role == 'FRM Cell Head':
            print("33333333333")
            query = """
            SELECT DISTINCT ON(o.unique_id)
                    o.unique_id,  -- Assuming this is the unique identifier for the alert
                    c.acctno,     -- Account number from customers table
                    c.custcd,     -- Customer ID from customers table
                    c.customername,  -- Customer name from customers table
                    o.branch,
                    o.query_number AS rule_number,
                    
                    o.comments_by_ro_checker,
                    o.ro_maker_files,
                    o.ro_checker_files,
                    o.date_updated_by_ro_checker,
                    o.comments_by_ro_maker,
                    o.date_updated_by_ro_maker,
                    o.comments_by_co_maker,
                    o.date_updated_by_co_maker,
                    o.comments_by_co_checker,
                    o.date_updated_by_co_checker,
                    o.co_maker_files,
                    o.co_checker_files,
                    o.co_maker_status,
                    o.co_maker_action,
                    o.co_checker_status,
                    o.co_checker_action,
                    o.reply_body,
                    o.attachments,
                    o.frm_cell_head,
                    o.date_updated_by_frm_cell_head,
                    o.frm_cell_head_files,
                    o.audit_comment,
                    o.audit_files,
                    o.date_updated_by_audit,
                    o.description,  -- Description from output table
                    o.date AS raised_date,  -- Date (Alert raised date) from output table
                    o.branch_region , -- Branch region from output table,
                    o.email_id,
                    o.frm_after_audit,
                    o.frm_after_audit_date,
                    o.frm_after_audit_files
            FROM output_table o
            JOIN user_details u ON o.branch_region = u.branch_region
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE frm_after_Audit is NOT NULL;


            """
        # elif user_role == 'CO Checker':
        #     query = """
        #     SELECT DISTINCT ON (o.unique_id)
        #         o.unique_id,  -- Assuming you want distinct results based on the unique_id
        #         c.acctno,     
        #         c.custcd, 
        #         c.customername, 
        #         o.branch,
        #         o.comments_by_ro_checker,
        #         o.ro_maker_files,
        #         o.ro_checker_files,
        #         o.date_updated_by_ro_checker,
        #         o.comments_by_ro_maker,
        #         o.date_updated_by_ro_maker,
        #         o.query_number AS rule_number, 
        #         o.frm_cell_head, 
        #         o.description,    
        #         o.*,              
        #         o.date AS raised_date,          
        #         o.branch_region,
        #         u., c.  -- or you can select specific columns from user_details and customers if needed
        #     FROM output_table o
        #     JOIN user_details u ON o.branch_region = u.branch_region
        #     LEFT JOIN customers c ON o.customer_id = c.custcd
            
        #     WHERE o.co_maker_scn_comments IS NOT NULL 
        #     AND (o.reply_body IS NOT NULL OR o.attachments IS NOT NULL);
        #     """
        
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")
        matched_data = [dict(zip(columns, row)) for row in results]
        # Debug print: Only show the 'date_updated_by_frm_cell_head' column
        if 'date_updated_by_frm_cell_head' in columns:
            updated_dates = [row.get('date_updated_by_frm_cell_head') for row in matched_data]
            print("Date Updated by FRM Cell Head:", updated_dates)
        else:
            print("The column 'date_updated_by_frm_cell_head' was not found in the query results.")

        # print("Matched Data:", matched_data)  # Debug print
        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Connection closed123.")




@app.route('/closed_alerts')
@jwt_required()
def closed_alerts():
    try:
        user_id = get_jwt_identity()

        # Fetch user details and validate user
        user = UserDetails.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        user_role = user.role

        # Database connection
        connection = psycopg2.connect(
            dbname=dbname,
            user=os.getenv('DB_USER'),
            password=os.getenv('DB_PASSWORD'),
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT')
        )
        cursor = connection.cursor()

        print("Connected to the database.")

        if user_role == 'CO Checker':
            # Query for CO Checker role: Fetch closed alerts where status is not recommended for RFA
            query = """
            SELECT DISTINCT ON (o.unique_id)
                o.unique_id,                  -- Unique identifier for the alert
                c.acctno,                     -- Account number from customers table
                c.custcd,                     -- Customer ID from customers table
                c.customername,               -- Customer name from customers table
                o.branch,                     -- Branch
                o.query_number AS rule_number, -- Rule number
                o.description,                -- Description of the alert
                o.co_checker_status,          -- Status for CO Checker
                o.co_checker_action,          -- Action taken by CO Checker
                o.date_updated_by_co_checker AS raised_date, -- Closed date (if applicable)
                o.branch_region     ,          -- Branch region
                o.*
            FROM output_table o
            LEFT JOIN customers c ON o.customer_id = c.custcd
            WHERE
            (o.co_checker_action = 'Close Alert') OR
            (o.co_checker_scn_action = 'Existing RFA Status removed(Alert Closed)')
            ORDER BY o.unique_id, o.date DESC; -- Ensure distinct results based on unique_id

            """
        else:
            return jsonify({"status": "error", "message": "Access denied for this role"}), 403

        # Execute query and fetch results
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        print(f"Fetched {len(results)} rows from the database.")

        # Map results to a list of dictionaries
        matched_data = [dict(zip(columns, row)) for row in results]

        return jsonify({"columns": columns, "data": matched_data})
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        # Close database connection
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("Database connection closed.")












if __name__ == '__main__':
    app.run(host="0.0.0.0",debug=True, port=5002)
